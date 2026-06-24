import os
import datetime
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.http import HttpResponse

from accounts.decorators import admin_required
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import inch

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Report
from analytics.models import DailyAdmission
from inventory.models import Inventory, Medicine
from predictions.models import Prediction, RiskPrediction
from disease_risk.models import DiseaseRiskPrediction


def generate_report_id(report_type):
    now = datetime.datetime.now()
    prefix = report_type[:3].upper()
    return f"{prefix}-{now.strftime('%Y%m%d%H%M%S')}"


def save_pdf_response(doc, buffer, request, report_type, title, description, date_start=None, date_end=None):
    path = doc.filename
    report_id = generate_report_id(report_type)
    Report.objects.create(
        report_id=report_id,
        report_type=report_type,
        title=title,
        description=description,
        generated_by=request.user,
        file_path=path,
        format='pdf',
        date_range_start=date_start,
        date_range_end=date_end,
    )
    messages.success(request, f'{title} generated successfully.')
    return redirect('reports:report_list')


@admin_required
def report_list(request):
    reports = Report.objects.all()
    return render(request, 'reports/report_list.html', {'reports': reports})


@admin_required
def generate_admission_report(request):
    admissions = DailyAdmission.objects.all().order_by('date')
    if not admissions.exists():
        messages.warning(request, 'No admission data available.')
        return redirect('reports:report_list')

    date_start = admissions.first().date
    date_end = admissions.last().date

    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"admission_report_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    filepath = os.path.join(reports_dir, filename)

    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()

    elements = []

    elements.append(Paragraph('Admission Report', styles['Title']))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"Date Range: {date_start} to {date_end}", styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))

    data = [['Date', 'Admissions', 'Discharges', 'Emergency Cases']]
    total_admissions = 0
    total_discharges = 0
    total_emergency = 0

    for a in admissions:
        data.append([str(a.date), a.total_admissions, a.total_discharges, a.emergency_cases])
        total_admissions += a.total_admissions
        total_discharges += a.total_discharges
        total_emergency += a.emergency_cases

    data.append(['Total', total_admissions, total_discharges, total_emergency])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4B8BBE')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E8F0FE')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))
    elements.append(Paragraph(f"Summary - Total Admissions: {total_admissions}, Discharges: {total_discharges}, Emergency Cases: {total_emergency}", styles['Normal']))

    doc.build(elements)
    return save_pdf_response(doc, None, request, 'admission', 'Admission Report',
                             f'Admission report from {date_start} to {date_end}',
                             date_start, date_end)


@admin_required
def generate_inventory_report(request):
    items = Inventory.objects.select_related('medicine').all().order_by('medicine__name')
    if not items.exists():
        messages.warning(request, 'No inventory data available.')
        return redirect('reports:report_list')

    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"inventory_report_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    filepath = os.path.join(reports_dir, filename)

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph('Inventory Report', styles['Title']))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))

    data = [['Medicine Name', 'Batch Number', 'Quantity', 'Expiry Date', 'Status']]
    for item in items:
        status = 'Expired' if item.is_expired() else ('Low Stock' if item.is_low_stock() else 'In Stock')
        data.append([item.medicine.name, item.batch_number, item.quantity, str(item.expiry_date), status])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4B8BBE')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(table)
    doc.build(elements)
    return save_pdf_response(doc, None, request, 'inventory', 'Inventory Report',
                             'Comprehensive inventory status report')


@admin_required
def generate_prediction_report(request):
    predictions = Prediction.objects.all().order_by('-predicted_date')
    risk_predictions = RiskPrediction.objects.all().order_by('-risk_score')

    if not predictions.exists() and not risk_predictions.exists():
        messages.warning(request, 'No prediction data available.')
        return redirect('reports:report_list')

    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"prediction_report_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    filepath = os.path.join(reports_dir, filename)

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph('Prediction Report', styles['Title']))
    elements.append(Spacer(1, 0.2 * inch))

    if predictions.exists():
        elements.append(Paragraph('Predictions', styles['Heading2']))
        elements.append(Spacer(1, 0.1 * inch))
        pred_data = [['Type', 'Date', 'Value', 'Confidence']]
        for p in predictions:
            pred_data.append([p.prediction_type, str(p.predicted_date), p.predicted_value, f"{p.confidence_score:.2f}%"])
        pred_table = Table(pred_data)
        pred_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4B8BBE')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(pred_table)
        elements.append(Spacer(1, 0.3 * inch))

    if risk_predictions.exists():
        elements.append(Paragraph('Risk Predictions', styles['Heading2']))
        elements.append(Spacer(1, 0.1 * inch))
        risk_data = [['Patient', 'Age', 'Condition', 'Risk Level', 'Risk Score']]
        for rp in risk_predictions:
            risk_data.append([rp.patient_name, rp.age, rp.predicted_condition, rp.risk_level.upper(), f"{rp.risk_score:.2f}%"])
        risk_table = Table(risk_data)
        risk_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC3545')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(risk_table)

    doc.build(elements)
    return save_pdf_response(doc, None, request, 'prediction', 'Prediction Report',
                             'Prediction and risk assessment report')


def export_excel_response(data, headers, filename_prefix):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    header_fill = PatternFill(start_color='4B8BBE', end_color='4B8BBE', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@admin_required
def export_admissions_excel(request):
    admissions = DailyAdmission.objects.all().order_by('date')
    if not admissions.exists():
        messages.warning(request, 'No admission data to export.')
        return redirect('reports:report_list')
    headers = ['Date', 'Admissions', 'Discharges', 'Emergency Cases']
    data = [[str(a.date), a.total_admissions, a.total_discharges, a.emergency_cases] for a in admissions]
    return export_excel_response(data, headers, 'admissions')


@admin_required
def export_inventory_excel(request):
    items = Inventory.objects.select_related('medicine').all().order_by('medicine__name')
    if not items.exists():
        messages.warning(request, 'No inventory data to export.')
        return redirect('reports:report_list')
    headers = ['Medicine Name', 'Batch Number', 'Quantity', 'Expiry Date', 'Status']
    data = []
    for item in items:
        status = 'Expired' if item.is_expired() else ('Low Stock' if item.is_low_stock() else 'In Stock')
        data.append([item.medicine.name, item.batch_number, item.quantity, str(item.expiry_date), status])
    return export_excel_response(data, headers, 'inventory')


@admin_required
def export_predictions_excel(request):
    predictions = Prediction.objects.all().order_by('-predicted_date')
    risk_predictions = RiskPrediction.objects.all().order_by('-risk_score')
    if not predictions.exists() and not risk_predictions.exists():
        messages.warning(request, 'No prediction data to export.')
        return redirect('reports:report_list')

    wb = openpyxl.Workbook()

    headers_pred = ['Type', 'Date', 'Value', 'Confidence']
    data_pred = [[p.prediction_type, str(p.predicted_date), p.predicted_value, p.confidence_score] for p in predictions]

    ws_pred = wb.active
    ws_pred.title = 'Predictions'
    header_fill = PatternFill(start_color='4B8BBE', end_color='4B8BBE', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    for col_idx, header in enumerate(headers_pred, 1):
        cell = ws_pred.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for row_idx, row_data in enumerate(data_pred, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_pred.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
    for col_idx in range(1, len(headers_pred) + 1):
        ws_pred.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    if risk_predictions.exists():
        ws_risk = wb.create_sheet('Risk Predictions')
        headers_risk = ['Patient', 'Age', 'Condition', 'Risk Level', 'Risk Score']
        data_risk = [[rp.patient_name, rp.age, rp.predicted_condition, rp.risk_level.upper(), rp.risk_score] for rp in risk_predictions]
        risk_fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
        for col_idx, header in enumerate(headers_risk, 1):
            cell = ws_risk.cell(row=1, column=col_idx, value=header)
            cell.fill = risk_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        for row_idx, row_data in enumerate(data_risk, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_risk.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='center')
        for col_idx in range(1, len(headers_risk) + 1):
            ws_risk.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="predictions_{datetime.datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
    wb.save(response)
    return response


@admin_required
def generate_disease_risk_report(request):
    predictions = DiseaseRiskPrediction.objects.all().order_by('-risk_score')

    if not predictions.exists():
        messages.warning(request, 'No disease risk prediction data available.')
        return redirect('reports:report_list')

    reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
    os.makedirs(reports_dir, exist_ok=True)

    filename = f"disease_risk_report_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    filepath = os.path.join(reports_dir, filename)

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph('Disease Risk Prediction Report', styles['Title']))
    elements.append(Spacer(1, 0.2 * inch))
    elements.append(Paragraph(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))

    total = predictions.count()
    high = predictions.filter(risk_level='high').count()
    medium = predictions.filter(risk_level='medium').count()
    low = predictions.filter(risk_level='low').count()

    elements.append(Paragraph(f'Summary: Total: {total} | High Risk: {high} | Medium: {medium} | Low: {low}', styles['Normal']))
    elements.append(Spacer(1, 0.3 * inch))

    risk_data = [['Patient Name', 'Disease', 'Risk Level', 'Risk Score', 'Confidence', 'Age', 'BP', 'Sugar', 'BMI', 'Chol']]
    for p in predictions:
        risk_data.append([
            p.patient_name, p.get_disease_type_display(), p.risk_level.upper(),
            f"{p.risk_score:.1f}%", f"{p.confidence_score:.1f}%",
            str(p.age), f"{p.blood_pressure}", f"{p.sugar_level}",
            f"{p.bmi}", f"{p.cholesterol}"
        ])

    table = Table(risk_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    elements.append(table)

    doc.build(elements)
    return save_pdf_response(doc, None, request, 'disease_risk', 'Disease Risk Report',
                             'Disease risk prediction report')


@admin_required
def export_disease_risk_excel(request):
    predictions = DiseaseRiskPrediction.objects.all().order_by('-risk_score')

    if not predictions.exists():
        messages.warning(request, 'No disease risk data to export.')
        return redirect('reports:report_list')

    headers = ['Patient Name', 'Disease', 'Risk Level', 'Risk Score', 'Confidence',
               'Age', 'Gender', 'BP', 'Sugar', 'BMI', 'Cholesterol', 'Heart Rate', 'Date']
    data = []
    for p in predictions:
        data.append([
            p.patient_name, p.get_disease_type_display(), p.risk_level.upper(),
            p.risk_score, p.confidence_score, p.age, p.get_gender_display(),
            p.blood_pressure, p.sugar_level, p.bmi, p.cholesterol, p.heart_rate,
            p.created_at.strftime('%Y-%m-%d')
        ])
    return export_excel_response(data, headers, 'disease_risk')
